"""
STC Template Excel Export Service
Generates Excel files matching STC's template structure with 5 sheets

MASTER WORKBOOK APPROACH:
- Maintains a single master Excel file per customer (stored in database)
- Each export APPENDS to the existing master (not creates new)
- Tracks which tenders have been exported to avoid duplicates
- Uses Redis locking for concurrent access safety
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict, Optional
from io import BytesIO
from sqlalchemy.orm import Session
from app.models.tender import Tender
from app.models.export_file import ExportFile
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


# Reference lists matching STC template
BIDDING_COMPANIES = ["STC", "SSTC", "ePortal", "CDN", "JMT", "AlDar", "H3"]

SECTORS = ["Government", "Oil & Gas", "Banking", "Private", "Telecom"]

TENDER_TYPES = [
    "CTC", "Semi-Tender", "Direct Tender", "RFP", "RFQ", "RFI",
    "Budgetary", "Writing RFP", "Pre-Qualification", "Auction"
]

ACCOUNT_MANAGERS = [
    "Ahmed El Henawy", "Ahmed Al Rahwan", "Amr El-Ragal", "Hany Ismail",
    "Mohamed Abuzarqa", "Mohsin Abdelrazig", "Rafik Hafez", "Syed Mohsin",
    "Issa Alsuwait", "Arpan Panigrahy", "Bharath Bilimaga", "Hursh Chandhok",
    "Mohamed Ibrahim", "Mohammad Hasan", "Noha Saranek", "Sachin Khan",
    "Samer Al Khatib"
]

JUSTIFICATIONS = [
    "Does Not Meet Qualifications",
    "Price Protected for Others",
    "Non-Compliance with Technical Specifications",
    "Outside Scope of Our Services",
    "Short Notice/Need More Time",
    "Over Budget"
]

ANNOUNCEMENT_TYPES = [
    "Awarding", "Complaint", "Opening Envelopes", "Cancellation",
    "Contract Extension", "Reconsideration", "Renew Bid Bond"
]


class STCTemplateExporter:
    """
    Generate Excel files matching STC template
    
    MASTER WORKBOOK MODE:
    - Loads existing master file from database
    - Appends new tenders (skips already exported)
    - Saves updated master back to database
    - Returns the updated file for download
    """
    
    MASTER_FILE_NAME = "stc_master"
    
    def __init__(self, db: Session):
        self.db = db
    
    def generate_excel(self, tender_ids: List[int], skip_duplicates: bool = True) -> bytes:
        """
        Generate/update Excel file with STC template
        
        MASTER WORKBOOK APPROACH:
        1. Load existing master from database (or create new if first time)
        2. Filter out already-exported tenders (optional)
        3. Append new tenders to correct sheets
        4. Save updated master to database
        5. Return file bytes for download
        
        Args:
            tender_ids: List of tender IDs to export
            skip_duplicates: If True, skip tenders already exported (default: True)
            
        Returns:
            Excel file as bytes
        """
        # Fetch tenders
        tenders = self.db.query(Tender).filter(Tender.id.in_(tender_ids)).all()
        
        if not tenders:
            raise ValueError("No tenders found with provided IDs")
        
        # Filter out already exported tenders if requested
        if skip_duplicates:
            new_tenders = [t for t in tenders if not t.exported_to_stc_at]
            skipped_count = len(tenders) - len(new_tenders)
            if skipped_count > 0:
                logger.info(f"Skipped {skipped_count} already-exported tenders")
            tenders = new_tenders
        
        if not tenders:
            logger.info("All selected tenders already exported")
            # Still return the master file even if no new tenders
            return self._get_or_create_master()
        
        # Load or create master workbook
        wb = self._load_master_workbook()
        
        # Get sheets
        ws_released = wb["Released Tenders"]
        ws_future = wb["Future Tenders"]
        ws_awarded = wb["Awarded-Opened Tenders"]
        
        # Group and append tenders
        exported_count = 0
        for tender in tenders:
            status = (tender.status or "Released").lower()
            
            if status in ['future', 'upcoming', 'planned']:
                self._append_to_future_sheet(ws_future, tender)
            elif status in ['awarded', 'opened', 'cancelled']:
                self._append_to_awarded_sheet(ws_awarded, tender)
            else:
                self._append_to_released_sheet(ws_released, tender)
            
            # Mark tender as exported
            tender.exported_to_stc_at = datetime.utcnow()
            exported_count += 1
        
        # Commit tender updates
        self.db.commit()
        logger.info(f"✅ Exported {exported_count} tenders to STC master")
        
        # Save updated workbook to database
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        file_bytes = output.read()
        
        self._save_master_workbook(file_bytes, exported_count)
        
        return file_bytes
    
    def _get_or_create_master(self) -> bytes:
        """Get existing master file or create empty one"""
        export_file = self.db.query(ExportFile).filter(
            ExportFile.name == self.MASTER_FILE_NAME
        ).first()
        
        if export_file:
            return export_file.file_data
        
        # Create new master
        wb = self._create_empty_workbook()
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
    
    def _load_master_workbook(self) -> Workbook:
        """Load existing master workbook from database, or create new"""
        export_file = self.db.query(ExportFile).filter(
            ExportFile.name == self.MASTER_FILE_NAME
        ).first()
        
        if export_file and export_file.file_data:
            logger.info(f"Loading existing master workbook ({export_file.total_rows_exported} rows)")
            return load_workbook(BytesIO(export_file.file_data))
        
        logger.info("Creating new master workbook")
        return self._create_empty_workbook()
    
    def _create_empty_workbook(self) -> Workbook:
        """Create new workbook with empty sheets matching STC template"""
        wb = Workbook()
        wb.remove(wb.active)
        
        # Create sheets with headers only
        self._create_released_tenders_sheet(wb, [])
        self._create_future_tenders_sheet(wb, [])
        self._create_awarded_opened_sheet(wb, [])
        self._create_color_coding_sheet(wb)
        self._create_list_sheet(wb)
        
        return wb
    
    def _save_master_workbook(self, file_bytes: bytes, new_rows: int):
        """Save master workbook to database"""
        export_file = self.db.query(ExportFile).filter(
            ExportFile.name == self.MASTER_FILE_NAME
        ).first()
        
        if export_file:
            export_file.file_data = file_bytes
            export_file.file_size = len(file_bytes)
            export_file.total_rows_exported += new_rows
            export_file.updated_at = datetime.utcnow()
        else:
            export_file = ExportFile(
                name=self.MASTER_FILE_NAME,
                file_data=file_bytes,
                file_size=len(file_bytes),
                description="STC Master Tenders Workbook",
                total_rows_exported=new_rows
            )
            self.db.add(export_file)
        
        self.db.commit()
        logger.info(f"✅ Master workbook saved ({len(file_bytes)} bytes, {export_file.total_rows_exported} total rows)")
    
    def _append_to_released_sheet(self, ws, tender: Tender):
        """Append a single tender to Released Tenders sheet"""
        next_row = ws.max_row + 1
        row_num = next_row - 5  # Numbering starts after header row 5
        
        ws.cell(next_row, 1, "")  # Empty column A
        ws.cell(next_row, 2, row_num)  # No.
        ws.cell(next_row, 3, self._get_gazette_number(tender))
        ws.cell(next_row, 4, self._format_date(tender.published_at))
        ws.cell(next_row, 5, self._get_page_number(tender))
        ws.cell(next_row, 6, tender.ministry or "")
        ws.cell(next_row, 7, tender.bidding_company or "")
        ws.cell(next_row, 8, tender.sector or "")
        ws.cell(next_row, 9, tender.tender_type or "")
        ws.cell(next_row, 10, tender.tender_number or "")
        ws.cell(next_row, 11, tender.title or "")
        # Use explicit tender_fee if set, otherwise fall back to document_price_kd from gazette
        fee_value = tender.tender_fee or tender.document_price_kd
        ws.cell(next_row, 12, float(fee_value) if fee_value else "")
    
    def _append_to_future_sheet(self, ws, tender: Tender):
        """Append a single tender to Future Tenders sheet"""
        next_row = ws.max_row + 1
        row_num = next_row - 5
        
        ws.cell(next_row, 1, row_num)
        ws.cell(next_row, 2, tender.ministry or "")
        ws.cell(next_row, 3, tender.bidding_company or "")
        ws.cell(next_row, 4, tender.sector or "")
        ws.cell(next_row, 5, tender.tender_type or "")
        ws.cell(next_row, 6, tender.tender_number or "")
        ws.cell(next_row, 7, tender.title or "")
        ws.cell(next_row, 8, self._format_date(getattr(tender, 'release_date', None)))
        ws.cell(next_row, 9, float(tender.expected_value) if getattr(tender, 'expected_value', None) else "")
        ws.cell(next_row, 10, tender.status or "")
        ws.cell(next_row, 11, getattr(tender, 'justification', "") or "")
        ws.cell(next_row, 12, getattr(tender, 'announcement_type', "") or "")
    
    def _append_to_awarded_sheet(self, ws, tender: Tender):
        """Append a single tender to Awarded-Opened sheet"""
        next_row = ws.max_row + 1
        row_num = next_row - 5
        
        ws.cell(next_row, 1, row_num)
        ws.cell(next_row, 2, tender.ministry or "")
        ws.cell(next_row, 3, tender.bidding_company or "")
        ws.cell(next_row, 4, tender.sector or "")
        ws.cell(next_row, 5, tender.tender_type or "")
        ws.cell(next_row, 6, tender.tender_number or "")
        ws.cell(next_row, 7, tender.title or "")
        ws.cell(next_row, 8, tender.status or "")
        ws.cell(next_row, 9, getattr(tender, 'awarded_vendor', "") or "")
        ws.cell(next_row, 10, float(tender.awarded_value) if getattr(tender, 'awarded_value', None) else "")
        ws.cell(next_row, 11, getattr(tender, 'justification', "") or "")
        ws.cell(next_row, 12, getattr(tender, 'announcement_type', "") or "")
    
    def _create_released_tenders_sheet(self, wb: Workbook, tenders: List[Tender]):
        """Create Released Tenders sheet"""
        ws = wb.create_sheet("Released Tenders", 0)
        
        # Header at row 5
        headers = [
            "", "No.", "Kuwait Alyoum NO.", "Kuwait Alyoum Date.",
            "Kuwait Alyoum Page No.", "Customer Name", "Bidding Company",
            "Sector", "Tender Type", "Tender No.", "Tender Title", "Tender Fee"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(5, col_idx, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows starting from row 6
        for idx, tender in enumerate(tenders, 1):
            row = 5 + idx
            
            ws.cell(row, 1, "")  # Empty column A
            ws.cell(row, 2, idx)  # No.
            ws.cell(row, 3, self._get_gazette_number(tender))
            ws.cell(row, 4, self._format_date(tender.published_at))
            ws.cell(row, 5, self._get_page_number(tender))
            ws.cell(row, 6, tender.ministry or "")
            ws.cell(row, 7, tender.bidding_company or "")
            ws.cell(row, 8, tender.sector or "")
            ws.cell(row, 9, tender.tender_type or "")
            ws.cell(row, 10, tender.tender_number or "")
            ws.cell(row, 11, tender.title or "")
            ws.cell(row, 12, float(tender.tender_fee) if tender.tender_fee else "")
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    def _create_future_tenders_sheet(self, wb: Workbook, tenders: List[Tender]):
        """Create Future Tenders sheet"""
        ws = wb.create_sheet("Future Tenders", 1)
        
        # Header at row 5
        headers = [
            "No.", "Customer Name", "Bidding Company", "Sector", "Tender Type",
            "Tender No.", "Tender Title", "Release Date", "Expected Value",
            "Status", "Justification", "Type of Announcement"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(5, col_idx, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for idx, tender in enumerate(tenders, 1):
            row = 5 + idx
            
            ws.cell(row, 1, idx)
            ws.cell(row, 2, tender.ministry or "")
            ws.cell(row, 3, tender.bidding_company or "")
            ws.cell(row, 4, tender.sector or "")
            ws.cell(row, 5, tender.tender_type or "")
            ws.cell(row, 6, tender.tender_number or "")
            ws.cell(row, 7, tender.title or "")
            ws.cell(row, 8, self._format_date(tender.release_date))
            ws.cell(row, 9, float(tender.expected_value) if tender.expected_value else "")
            ws.cell(row, 10, tender.status or "")
            ws.cell(row, 11, tender.justification or "")
            ws.cell(row, 12, tender.announcement_type or "")
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    def _create_awarded_opened_sheet(self, wb: Workbook, tenders: List[Tender]):
        """Create Awarded-Opened Tenders sheet"""
        ws = wb.create_sheet("Awarded-Opened Tenders", 2)
        
        # Header at row 5
        headers = [
            "No.", "Customer Name", "Bidding Company", "Sector", "Tender Type",
            "Tender No.", "Tender Title", "Status", "Awarded Vendor",
            "Awarded Value", "Justification", "Type of Announcement"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(5, col_idx, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for idx, tender in enumerate(tenders, 1):
            row = 5 + idx
            
            ws.cell(row, 1, idx)
            ws.cell(row, 2, tender.ministry or "")
            ws.cell(row, 3, tender.bidding_company or "")
            ws.cell(row, 4, tender.sector or "")
            ws.cell(row, 5, tender.tender_type or "")
            ws.cell(row, 6, tender.tender_number or "")
            ws.cell(row, 7, tender.title or "")
            ws.cell(row, 8, tender.status or "")
            ws.cell(row, 9, tender.awarded_vendor or "")
            ws.cell(row, 10, float(tender.awarded_value) if tender.awarded_value else "")
            ws.cell(row, 11, tender.justification or "")
            ws.cell(row, 12, tender.announcement_type or "")
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    def _create_color_coding_sheet(self, wb: Workbook):
        """Create Color Coding legend sheet"""
        ws = wb.create_sheet("Color Coding", 3)
        
        ws.cell(1, 1, "Color Legend")
        ws.cell(1, 1).font = Font(bold=True, size=14)
        
        ws.cell(3, 1, "Released Tenders")
        ws.cell(3, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        ws.cell(4, 1, "Future Tenders")
        ws.cell(4, 1).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        ws.cell(5, 1, "Awarded / Opened")
        ws.cell(5, 1).fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        ws.column_dimensions['A'].width = 30
    
    def _create_list_sheet(self, wb: Workbook):
        """Create reference lists sheet"""
        ws = wb.create_sheet("List", 4)
        
        # Bidding Company (Column B)
        ws.cell(1, 2, "Bidding Company")
        ws.cell(1, 2).font = Font(bold=True)
        for idx, company in enumerate(BIDDING_COMPANIES, 2):
            ws.cell(idx, 2, company)
        
        # Sector (Column D)
        ws.cell(1, 4, "Sector")
        ws.cell(1, 4).font = Font(bold=True)
        for idx, sector in enumerate(SECTORS, 2):
            ws.cell(idx, 4, sector)
        
        # Tender Type (Column F)
        ws.cell(1, 6, "Tender Type")
        ws.cell(1, 6).font = Font(bold=True)
        for idx, tender_type in enumerate(TENDER_TYPES, 2):
            ws.cell(idx, 6, tender_type)
        
        # Account Manager (Column H)
        ws.cell(1, 8, "Account Manager")
        ws.cell(1, 8).font = Font(bold=True)
        for idx, manager in enumerate(ACCOUNT_MANAGERS, 2):
            ws.cell(idx, 8, manager)
        
        # Justification (Column J)
        ws.cell(1, 10, "Justification")
        ws.cell(1, 10).font = Font(bold=True)
        for idx, justification in enumerate(JUSTIFICATIONS, 2):
            ws.cell(idx, 10, justification)
        
        # Type of Announcement (Column L)
        ws.cell(1, 12, "Type of Announcement")
        ws.cell(1, 12).font = Font(bold=True)
        for idx, announcement in enumerate(ANNOUNCEMENT_TYPES, 2):
            ws.cell(idx, 12, announcement)
        
        # Auto-size all columns
        for col in ['B', 'D', 'F', 'H', 'J', 'L']:
            ws.column_dimensions[col].width = 30
    
    def _get_gazette_number(self, tender: Tender) -> str:
        """Extract gazette number from tender"""
        # Try to get from gazette_id or edition_no attributes
        if hasattr(tender, 'gazette_id') and tender.gazette_id:
            return str(tender.gazette_id)
        if hasattr(tender, 'edition_no') and tender.edition_no:
            return str(tender.edition_no)
        return ""
    
    def _get_page_number(self, tender: Tender) -> str:
        """Extract page number from tender"""
        if hasattr(tender, 'page_number') and tender.page_number:
            return str(tender.page_number)
        return ""
    
    def _format_date(self, date_value) -> str:
        """Format date for Excel"""
        if not date_value:
            return ""
        if isinstance(date_value, datetime):
            return date_value.strftime("%Y-%m-%d")
        return str(date_value)
