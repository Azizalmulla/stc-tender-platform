"""
Generic Tender Excel Export Service

Produces a clean, neutral Excel report for any company.
No STC-specific branding or internal fields.
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from app.models.tender import Tender
from datetime import datetime, timezone
import io
import logging

logger = logging.getLogger(__name__)

# Colors
COLOR_HEADER_BG = "1F3864"   # Dark navy
COLOR_HEADER_FG = "FFFFFF"   # White
COLOR_ROW_ALT   = "EBF0FA"   # Light blue-grey alternate row
COLOR_ROW_EVEN  = "FFFFFF"   # White
COLOR_ACCENT    = "2E74B5"   # Blue accent for title row
COLOR_BORDER    = "BDD0E9"

CATEGORY_LABELS = {
    "tenders":   "مناقصة / Tender",
    "auctions":  "مزايدة / Auction",
    "practices": "ممارسة / Practice",
}

COLUMNS = [
    ("No.",            8),
    ("Tender No.",     18),
    ("Ministry (AR)",  28),
    ("Category",       18),
    ("Sector",         20),
    ("Summary (EN)",   55),
    ("Summary (AR)",   55),
    ("Deadline",       18),
    ("Meeting Date",   18),
    ("Doc. Price (KD)", 16),
    ("Est. Value (KD)", 16),
    ("Published",      14),
    ("Source Link",    40),
]


def _thin_border():
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_font():
    return Font(name="Calibri", bold=True, color=COLOR_HEADER_FG, size=11)


def _cell_font(bold=False):
    return Font(name="Calibri", bold=bold, size=10)


def _wrap():
    return Alignment(wrap_text=True, vertical="top")


class GenericTenderExporter:
    def __init__(self, db: Session):
        self.db = db

    def generate_excel(
        self,
        days_back: int = 7,
        category: str = None,
        sector: str = None,
    ) -> bytes:
        """
        Generate a generic Excel report of recent tenders.

        Args:
            days_back: How many days back to include (default 7)
            category: Optional filter by category (tenders/auctions/practices)
            sector: Optional filter by sector tag
        """
        tenders = self._fetch_tenders(days_back, category, sector)
        logger.info(f"Exporting {len(tenders)} tenders to generic Excel")

        wb = Workbook()
        ws = wb.active
        ws.title = "Kuwait Tech Tenders"

        self._write_title_row(ws, days_back, len(tenders))
        self._write_header_row(ws)
        self._write_data_rows(ws, tenders)
        self._apply_column_widths(ws)
        self._freeze_panes(ws)

        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        self._write_summary_sheet(ws_summary, tenders, days_back)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _fetch_tenders(self, days_back, category, sector):
        from datetime import timedelta
        cutoff = datetime.now(timezone.utc) - timedelta(days=days_back)
        q = self.db.query(Tender).filter(Tender.created_at >= cutoff)
        if category:
            q = q.filter(Tender.category == category)
        if sector:
            q = q.filter(Tender.ai_sectors.any(sector))
        return q.order_by(Tender.deadline.asc().nullslast()).all()

    def _write_title_row(self, ws, days_back, count):
        ws.merge_cells("A1:M1")
        cell = ws["A1"]
        now_str = datetime.now().strftime("%B %d, %Y")
        cell.value = f"Kuwait Technology Tenders Report  |  Last {days_back} Days  |  {count} Tenders  |  Generated: {now_str}"
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
        cell.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

    def _write_header_row(self, ws):
        for col_idx, (label, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=2, column=col_idx, value=label)
            cell.font = _header_font()
            cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _thin_border()
        ws.row_dimensions[2].height = 22

    def _write_data_rows(self, ws, tenders):
        for idx, t in enumerate(tenders, start=1):
            row = idx + 2  # rows 3+
            fill_color = COLOR_ROW_ALT if idx % 2 == 0 else COLOR_ROW_EVEN

            sectors_str = ", ".join(t.ai_sectors) if t.ai_sectors else ""
            deadline_str = t.deadline.strftime("%Y-%m-%d") if t.deadline else ""
            meeting_str = t.meeting_date.strftime("%Y-%m-%d %H:%M") if t.meeting_date else ""
            published_str = t.published_at.strftime("%Y-%m-%d") if t.published_at else ""
            category_str = CATEGORY_LABELS.get(t.category, t.category or "")
            doc_price = float(t.document_price_kd) if t.document_price_kd else ""
            est_value = float(t.expected_value) if t.expected_value else ""

            values = [
                idx,
                t.tender_number or "",
                t.ministry or "",
                category_str,
                sectors_str,
                t.summary_en or "",
                t.summary_ar or "",
                deadline_str,
                meeting_str,
                doc_price,
                est_value,
                published_str,
                t.url or "",
            ]

            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = _cell_font()
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.alignment = _wrap()
                cell.border = _thin_border()
                # Right-align numbers
                if col_idx in (1, 10, 11):
                    cell.alignment = Alignment(horizontal="right", vertical="top")
                # Make source link blue
                if col_idx == 13 and val:
                    cell.font = Font(name="Calibri", size=10, color="2E74B5", underline="single")

            ws.row_dimensions[row].height = 60

    def _apply_column_widths(self, ws):
        for col_idx, (_, width) in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _freeze_panes(self, ws):
        ws.freeze_panes = "A3"

    def _write_summary_sheet(self, ws, tenders, days_back):
        ws["A1"] = "Summary"
        ws["A1"].font = Font(bold=True, size=14, color=COLOR_ACCENT)

        rows = [
            ("Total Tenders", len(tenders)),
            ("Days Covered", days_back),
            ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("", ""),
            ("By Category", ""),
        ]

        from collections import Counter
        cat_counts = Counter(t.category for t in tenders)
        for cat, cnt in cat_counts.items():
            rows.append((f"  {CATEGORY_LABELS.get(cat, cat)}", cnt))

        rows.append(("", ""))
        rows.append(("By Sector", ""))
        sector_counts = Counter(s for t in tenders if t.ai_sectors for s in t.ai_sectors)
        for sector, cnt in sector_counts.most_common():
            rows.append((f"  {sector}", cnt))

        rows.append(("", ""))
        rows.append(("Upcoming Deadlines (next 7 days)", ""))
        from datetime import timedelta
        soon = datetime.now(timezone.utc) + timedelta(days=7)
        urgent = [t for t in tenders if t.deadline and t.deadline <= soon]
        for t in urgent:
            rows.append((f"  {t.ministry or 'Unknown'}", t.deadline.strftime("%Y-%m-%d") if t.deadline else ""))

        for r_idx, (label, val) in enumerate(rows, start=3):
            ws.cell(row=r_idx, column=1, value=label).font = Font(name="Calibri", size=10, bold=(val == ""))
            ws.cell(row=r_idx, column=2, value=val).font = Font(name="Calibri", size=10)

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20
